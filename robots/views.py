from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook

from .models import Robot


def generate_exel_report(request):
    last_week = timezone.now() - timezone.timedelta(days=7)
    robots = Robot.objects.filter(created__gte=last_week)

    wb = Workbook()

    models = robots.values_list('model', flat=True).distinct()

    for model in models:
        ws = wb.create_sheet(title=model)
        ws.append(['Модель', 'Версия', 'Количество за неделю'])

        model_data = robots.filter(model=model).values('version').annotate(count=Count('version'))

        for data in model_data:
            ws.append([model, data['version'], data['count']])

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robot_report.xlsx'
    wb.save(response)

    return response
